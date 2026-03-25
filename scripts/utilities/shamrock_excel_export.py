"""
SHAMROCK Excel Trade Log Exporter
Reads paper_trades.json + paper_portfolio.json and creates a
formatted Excel workbook: shamrock_trades.xlsx
"""
import json
import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

# ── Colour palette (matches StockGuru UI) ────────────────────────────────────
C_BG        = "002D34"   # dark teal background
C_ACCENT    = "00C4CC"   # cyan accent
C_GREEN     = "00E676"   # profit green
C_RED       = "FF3D71"   # loss red
C_GOLD      = "FFD700"   # warning gold
C_HEADER_FG = "FFFFFF"
C_HEADER_BG = "0A3D2A"
C_ROW_ALT   = "0D2D20"
C_ROW_EVEN  = "0A2218"

THIN = Side(border_style="thin", color="1A4A3A")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _hfill(color): return PatternFill("solid", fgColor=color)
def _font(bold=False, color="FFFFFF", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def _center(): return Alignment(horizontal="center", vertical="center")
def _left():   return Alignment(horizontal="left",   vertical="center")


def _load(path, default):
    try:
        if os.path.exists(path):
            with open(path) as f:
                return json.load(f)
    except Exception:
        pass
    return default


def _col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


def _header_row(ws, row, labels, fill_color=C_HEADER_BG):
    for c, label in enumerate(labels, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = _font(bold=True, size=10)
        cell.fill = _hfill(fill_color)
        cell.alignment = _center()
        cell.border = BORDER


def build_excel(output_path="shamrock_trades.xlsx"):
    data_dir = os.path.join(os.path.dirname(__file__), "data")
    trades    = _load(os.path.join(data_dir, "paper_trades.json"),    [])
    portfolio = _load(os.path.join(data_dir, "paper_portfolio.json"), {})

    wb = Workbook()

    # ── Sheet 1: Trade Log ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Trade Log"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 36

    # Title banner
    ws.merge_cells("A1:I1")
    title = ws["A1"]
    title.value = f"🍀  SHAMROCK Trade Log — Generated {datetime.datetime.now().strftime('%d %b %Y  %H:%M')}"
    title.font  = _font(bold=True, size=14, color=C_ACCENT)
    title.fill  = _hfill(C_BG)
    title.alignment = _center()

    headers = ["#", "Symbol", "Action", "Entry ₹", "Target ₹", "SL ₹", "Qty", "Score", "Status", "PnL ₹", "Timestamp"]
    _header_row(ws, 2, headers)

    col_widths = [4, 14, 8, 12, 12, 12, 6, 7, 12, 12, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for i, trade in enumerate(trades, 1):
        row = i + 2
        fill = _hfill(C_ROW_ALT if i % 2 == 0 else C_ROW_EVEN)
        action  = trade.get("action", "BUY")
        pnl     = float(trade.get("pnl", 0))
        score   = trade.get("score", 0)
        values  = [
            i,
            trade.get("symbol", "-"),
            action,
            trade.get("entry_price", 0),
            trade.get("target", 0),
            trade.get("stop_loss", 0),
            trade.get("qty", 0),
            score,
            trade.get("status", "SIM"),
            round(pnl, 2),
            trade.get("timestamp", "")[:19] if trade.get("timestamp") else "",
        ]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill   = fill
            cell.border = BORDER
            cell.alignment = _center()
            # Colour overrides
            if c == 3:   # Action
                cell.font = _font(bold=True, color=C_GREEN if action == "BUY" else C_RED)
            elif c == 8:  # Score
                color = C_GREEN if score >= 88 else C_GOLD if score >= 75 else C_RED
                cell.font = _font(color=color, bold=True)
            elif c == 10:  # PnL
                cell.font = _font(bold=True, color=C_GREEN if pnl > 0 else C_RED if pnl < 0 else "AAAAAA")
                cell.number_format = '₹#,##0.00'
            elif c in (4, 5, 6):
                cell.number_format = '₹#,##0.00'
            else:
                cell.font = _font(color="CCCCCC")

    # Summary row
    n_trades = len(trades)
    if n_trades:
        sum_row = n_trades + 3
        ws.merge_cells(f"A{sum_row}:I{sum_row}")
        total_pnl = sum(float(t.get("pnl", 0)) for t in trades)
        sum_cell = ws.cell(row=sum_row, column=1,
                           value=f"Total Trades: {n_trades}   |   Net PnL: ₹{total_pnl:,.2f}")
        sum_cell.font  = _font(bold=True, size=11, color=C_GREEN if total_pnl >= 0 else C_RED)
        sum_cell.fill  = _hfill(C_HEADER_BG)
        sum_cell.alignment = _center()
        pnl_cell = ws.cell(row=sum_row, column=10, value=round(total_pnl, 2))
        pnl_cell.font = _font(bold=True, color=C_GREEN if total_pnl >= 0 else C_RED)
        pnl_cell.fill = _hfill(C_HEADER_BG)
        pnl_cell.number_format = '₹#,##0.00'

    ws.freeze_panes = "A3"

    # ── Sheet 2: Portfolio Summary ────────────────────────────────────────────
    ws2 = wb.create_sheet("Portfolio")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:D1")
    hdr = ws2["A1"]
    hdr.value = "🍀  SHAMROCK Portfolio Snapshot"
    hdr.font  = _font(bold=True, size=13, color=C_ACCENT)
    hdr.fill  = _hfill(C_BG)
    hdr.alignment = _center()
    ws2.row_dimensions[1].height = 30

    fields = [
        ("Initial Capital",    f"₹{portfolio.get('capital', 500000):,.2f}"),
        ("Available Cash",     f"₹{portfolio.get('cash',    portfolio.get('capital', 500000)):,.2f}"),
        ("Open Positions",     len(portfolio.get("positions", {}))),
        ("Total Trades",       len(trades)),
        ("Winning Trades",     sum(1 for t in trades if float(t.get("pnl", 0)) > 0)),
        ("Losing Trades",      sum(1 for t in trades if float(t.get("pnl", 0)) < 0)),
        ("Total PnL",          f"₹{sum(float(t.get('pnl',0)) for t in trades):,.2f}"),
        ("Win Rate",           f"{(sum(1 for t in trades if float(t.get('pnl',0))>0)/max(len(trades),1)*100):.1f}%"),
        ("Generated At",       datetime.datetime.now().strftime("%d %b %Y %H:%M:%S")),
    ]
    for r, (label, value) in enumerate(fields, 2):
        lc = ws2.cell(row=r, column=1, value=label)
        vc = ws2.cell(row=r, column=2, value=str(value))
        lc.font = _font(bold=True, color="AAAAAA")
        vc.font = _font(color=C_GREEN if "₹" in str(value) else "FFFFFF")
        fill = _hfill(C_ROW_ALT if r % 2 == 0 else C_ROW_EVEN)
        for cell in [lc, vc]:
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = _left()
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 22

    # ── Sheet 3: Gate Analysis ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Gate Analysis")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:E1")
    g_hdr = ws3["A1"]
    g_hdr.value = "🍀  8-Gate Conviction System"
    g_hdr.font  = _font(bold=True, size=13, color=C_ACCENT)
    g_hdr.fill  = _hfill(C_BG)
    g_hdr.alignment = _center()
    ws3.row_dimensions[1].height = 30

    gate_defs = [
        ("Gate 1", "RSI Zone",    "RSI between 35–68",       "35 ≤ RSI ≤ 68"),
        ("Gate 2", "Volume",      "Volume ≥ 1.3× 20d avg",   "Vol ≥ 1.3×"),
        ("Gate 3", "EMA Trend",   "Price above EMA 50",      "P > EMA50"),
        ("Gate 4", "MACD",        "MACD bullish crossover",  "MACD > Signal"),
        ("Gate 5", "News Gate",   "No negative news",        "Sentiment ≥ 0"),
        ("Gate 6", "FII Gate",    "FII buying",              "FII Net > 0"),
        ("Gate 7", "PCR Gate",    "Put-Call ratio < 1.2",    "PCR < 1.2"),
        ("Gate 8", "Score Gate",  "Agent score ≥ 88",        "Score ≥ 88"),
    ]
    _header_row(ws3, 2, ["Gate", "Name", "Condition", "Threshold", "Required?"])
    for r, (gate, name, cond, thresh) in enumerate(gate_defs, 3):
        for c, val in enumerate([gate, name, cond, thresh, "✓ (6/8 minimum)"], 1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.fill = _hfill(C_ROW_ALT if r % 2 == 0 else C_ROW_EVEN)
            cell.border = BORDER
            cell.font = _font(color="CCCCCC" if c > 1 else C_ACCENT, bold=(c == 1))
            cell.alignment = _center()
    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 30
    ws3.column_dimensions["D"].width = 16
    ws3.column_dimensions["E"].width = 20

    ws3.merge_cells("A12:E12")
    note = ws3["A12"]
    note.value = "⚡ Minimum 6 of 8 gates must PASS for a paper trade to execute"
    note.font  = _font(bold=True, color=C_GOLD, size=11)
    note.fill  = _hfill(C_HEADER_BG)
    note.alignment = _center()

    wb.save(output_path)
    print(f"✅ SHAMROCK Excel export saved: {output_path}")
    return output_path


if __name__ == "__main__":
    out = build_excel(os.path.join(os.path.dirname(__file__), "shamrock_trades.xlsx"))
    print(f"File: {out}")
